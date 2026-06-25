from pdf2image import convert_from_path
from docx import Document
import markdown
import csv
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from .base import BasePreviewGenerator


class PDFPreviewGenerator(BasePreviewGenerator):
    """Generate preview from PDF first page"""

    def extract_content(self):
        return convert_from_path(self.file_path, first_page=1, last_page=1)[0]

    def create_image(self, content):
        return content  # PDF already returns an image


class DocxPreviewGenerator(BasePreviewGenerator):
    """Generate preview from DOCX document"""

    def extract_content(self):
        doc = Document(self.file_path)
        return "\n".join(p.text for p in doc.paragraphs[:8])


class XlsxPreviewGenerator(BasePreviewGenerator):
    """Generate visual preview from Excel spreadsheet"""

    CELL_WIDTH = 100
    CELL_HEIGHT = 25
    FONT_SIZE = 11
    GRID_COLOR = (200, 200, 200)
    HEADER_BG = (242, 242, 242)
    TEXT_COLOR = (0, 0, 0)
    BG_COLOR = (255, 255, 255)
    PADDING = 5

    def extract_content(self):
        """Extract spreadsheet data with formatting"""
        wb = load_workbook(self.file_path, data_only=True)
        sheet = wb.active

        data = []
        max_row = min(20, sheet.max_row or 20)
        max_col = min(10, sheet.max_column or 10)

        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)

                value = str(cell.value) if cell.value is not None else ""

                bg_color = self.BG_COLOR
                if cell.fill and cell.fill.start_color:
                    rgb = cell.fill.start_color.rgb
                    if rgb and isinstance(rgb, str) and len(rgb) >= 6:
                        if rgb not in ['00000000', 'FF000000', '00']:
                            try:
                                if len(rgb) == 8:
                                    rgb = rgb[2:]
                                bg_color = tuple(int(rgb[i:i + 2], 16) for i in (0, 2, 4))
                            except:
                                pass

                text_color = self.TEXT_COLOR
                if cell.font and cell.font.color:
                    rgb = cell.font.color.rgb
                    if rgb and isinstance(rgb, str) and len(rgb) >= 6:
                        if rgb not in ['00000000', 'FF000000', '00']:
                            try:
                                if len(rgb) == 8:
                                    rgb = rgb[2:]
                                text_color = tuple(int(rgb[i:i + 2], 16) for i in (0, 2, 4))
                            except:
                                pass

                is_bold = cell.font.bold if cell.font else False

                row_data.append({
                    'value': value,
                    'bg_color': bg_color,
                    'text_color': text_color,
                    'bold': is_bold
                })

            data.append(row_data)

        return data

    def create_image(self, content):
        """Create visual representation of spreadsheet"""
        if not content:
            return None

        rows = len(content)
        cols = len(content[0]) if content else 0

        width = cols * self.CELL_WIDTH + 1
        height = rows * self.CELL_HEIGHT + 1

        img = Image.new('RGB', (width, height), self.BG_COLOR)
        draw = ImageDraw.Draw(img)

        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", self.FONT_SIZE)
        except:
            try:
                font = ImageFont.truetype("arial.ttf", self.FONT_SIZE)
            except:
                font = ImageFont.load_default()

        for row_idx, row in enumerate(content):
            for col_idx, cell_data in enumerate(row):
                x = col_idx * self.CELL_WIDTH
                y = row_idx * self.CELL_HEIGHT

                bg_color = cell_data.get('bg_color', self.BG_COLOR)

                if row_idx == 0:
                    bg_color = self.HEADER_BG

                draw.rectangle(
                    [x, y, x + self.CELL_WIDTH, y + self.CELL_HEIGHT],
                    fill=bg_color,
                    outline=self.GRID_COLOR
                )

                value = cell_data['value']
                if value:
                    if len(value) > 15:
                        value = value[:12] + "..."

                    text_color = cell_data.get('text_color', self.TEXT_COLOR)

                    try:
                        bbox = draw.textbbox((0, 0), value, font=font)
                        text_height = bbox[3] - bbox[1]
                    except:
                        text_height = self.FONT_SIZE

                    text_x = x + self.PADDING
                    text_y = y + (self.CELL_HEIGHT - text_height) // 2

                    draw.text(
                        (text_x, text_y),
                        value,
                        fill=text_color,
                        font=font
                    )

        return img


class MarkdownPreviewGenerator(BasePreviewGenerator):
    """Generate preview from Markdown file"""

    def extract_content(self):
        with open(self.file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Convert markdown to plain text (strip HTML tags)
        html = markdown.markdown(content)
        text = html.replace("<", "").replace(">", "")
        return text


class CSVPreviewGenerator(BasePreviewGenerator):
    """Generate preview from CSV file"""

    def extract_content(self):
        lines = []
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 10:  # Limit to first 10 rows
                        break
                    line = " | ".join(str(cell) for cell in row[:5])  # First 5 columns
                    lines.append(line)
        except Exception as e:
            lines.append(f"Error reading CSV: {str(e)}")

        return "\n".join(lines)


class TxtPreviewGenerator(BasePreviewGenerator):
    """Generate preview from plain text file"""

    def extract_content(self):
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # Try with a different encoding if UTF-8 fails
            with open(self.file_path, 'r', encoding='latin-1') as f:
                return f.read()
