"""Markdown extraction functionality - Converts Excel files to Markdown format"""

import io
import logging
import os
import re
from typing import List, Tuple
import pandas as pd
from tabulate import tabulate

logger = logging.getLogger('django')

try:
    import openpyxl
    HAS_OPENPYXL = True
    logger.info("openpyxl available for enhanced Excel processing")
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not available. Excel hyperlink extraction will be limited.")


class MarkdownExtractor:
    """Handles Excel to Markdown conversion for better LLM processing"""

    def __init__(self, subdoc_max_chars: int = 500):
        self.subdoc_max_chars = subdoc_max_chars

    @staticmethod
    def sanitize_cell_content(df):
        """
        Cleans up cell content by converting internal line breaks to HTML <br> tags.
        This step runs BEFORE tabulate.
        """
        def replace_breaks(text):
            if isinstance(text, str):
                # Condense internal whitespace and clean up line breaks for <br> conversion
                text = re.sub(r'\s+', ' ', text)
                text = text.replace('\r\n', '<br>').replace('\n', '<br>').replace('\r', '<br>')
                return text.strip()
            return text

        return df.map(replace_breaks)

    @staticmethod
    def post_process_markdown(markdown_text):
        """
        Performs final cleanup on the generated Markdown text by:
        1. Reducing repetitive hyphens within general text.
        2. Standardizing the machine-generated table separator line.
        3. Condensing multiple spaces.
        """
        lines = markdown_text.splitlines()
        cleaned_lines = []

        for line in lines:
            # 1. Condense multiple hyphens (more than 3 repeating hyphens to 1 hyphen)
            line = re.sub(r'-{4,}', '-', line)

            # 2. Check for the machine-generated separator line and standardize it.
            if re.search(r'^\s*\|[:\s\-]+\|', line) and not re.search(r'[a-zA-Z0-9<>]', line):
                # Count the number of delimiters (pipes) to determine column count.
                num_columns = line.count('|') - 1
                if num_columns > 0:
                    # Create the standard separator: |---|---|...|
                    standard_separator = '|' + ':---|' * num_columns
                    cleaned_lines.append(standard_separator)
                    continue

            # 3. Condense multiple whitespace characters
            line = re.sub(r'\s{2,}', ' ', line)

            # Add the cleaned line
            cleaned_lines.append(line.strip())

        # 4. Remove any empty lines created by the filtering process
        final_output = '\n'.join([line for line in cleaned_lines if line])

        # Condense multiple blank lines
        final_output = re.sub(r'\n{3,}', '\n\n', final_output)

        return final_output.strip()

    def spreadsheet_to_markdown(self, content_bytes: bytes, filename: str = "spreadsheet") -> str:
        """
        Convert Excel/CSV spreadsheet to Markdown format

        Args:
            content_bytes: Spreadsheet file content as bytes
            filename: Name of the file for header

        Returns:
            Markdown formatted text
        """
        file_extension = os.path.splitext(filename)[1].lower()
        markdown_output = f"# {filename}\n\n"

        try:
            if file_extension == '.csv':
                # Read CSV without header assumption
                df = pd.read_csv(io.BytesIO(content_bytes), header=None, keep_default_na=False)
                df.columns = df.columns.astype(str)
                df = self.sanitize_cell_content(df)

                markdown_output += f"## {os.path.basename(filename).replace(file_extension, '')}\n\n"
                markdown_output += tabulate(df, headers='keys', tablefmt='pipe', showindex=False)
                markdown_output += "\n\n---\n"

            elif file_extension in ['.xlsx', '.xls']:
                xls = pd.ExcelFile(io.BytesIO(content_bytes))
                for sheet_name in xls.sheet_names:
                    # Read all content without assuming a header row
                    df = xls.parse(sheet_name, header=None, keep_default_na=False)

                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.warning(f"Sheet '{sheet_name}' is empty, skipping...")
                        continue

                    # Apply cleanup and conversion
                    df.columns = df.columns.astype(str)
                    df = self.sanitize_cell_content(df)

                    markdown_output += f"## {sheet_name}\n\n"
                    markdown_output += tabulate(df, headers='keys', tablefmt='pipe', showindex=False)
                    markdown_output += "\n\n---\n"

            else:
                return f"Unsupported file format: {file_extension}"

            return self.post_process_markdown(markdown_output)

        except Exception as e:
            logger.error(f"Error converting spreadsheet to markdown: {e}")
            return f"Error processing file: {e}"

    def extract_limited_content(self, content_bytes: bytes, max_chars: int = None, filename: str = "spreadsheet") -> str:
        """
        Extract limited content from Excel file in Markdown format for LLM processing

        Args:
            content_bytes: Excel file content as bytes
            max_chars: Maximum characters to extract
            filename: Name of the file

        Returns:
            Limited Markdown formatted text for LLM processing
        """
        if max_chars is None:
            max_chars = self.subdoc_max_chars

        try:
            # Convert to markdown
            markdown_content = self.spreadsheet_to_markdown(content_bytes, filename)

            logger.info("=" * 80)
            logger.info("EXCEL TO MARKDOWN CONVERSION COMPLETE")
            logger.info(f"Total content: {len(markdown_content)} characters")
            logger.info("=" * 80)

            # Apply character limit if needed
            if len(markdown_content) > max_chars:
                lines = markdown_content.split('\n')
                truncated_text = []
                current_length = 0

                for line in lines:
                    if current_length + len(line) + 1 > max_chars - 50:
                        break
                    truncated_text.append(line)
                    current_length += len(line) + 1

                markdown_content = '\n'.join(truncated_text) + "\n\n[Content truncated]"
                logger.info(f"Markdown content truncated to {len(markdown_content)} chars")

            # Log sample content
            logger.info("=" * 80)
            logger.info("MARKDOWN CONTENT BEING SENT TO LLM:")
            logger.info("=" * 80)
            logger.info(markdown_content[:1000] if len(markdown_content) > 1000 else markdown_content)
            if len(markdown_content) > 1000:
                logger.info(f"... [TRUNCATED - FULL CONTENT IS {len(markdown_content)} CHARS] ...")
            logger.info("=" * 80)

            return markdown_content

        except Exception as e:
            logger.error(f"Error extracting Excel content as Markdown: {e}")
            return ""

    def extract_comprehensive_content_for_urls(self, content_bytes: bytes, filename: str = "spreadsheet") -> Tuple[str, List[str]]:
        """
        Extract COMPLETE Excel content in Markdown format and hyperlinks using openpyxl

        Args:
            content_bytes: Excel file content as bytes
            filename: Name of the file

        Returns:
            Tuple of (comprehensive_markdown_text, extracted_urls)
        """
        try:
            logger.info("=" * 80)
            logger.info("EXTRACTING COMPREHENSIVE EXCEL CONTENT AS MARKDOWN WITH URL EXTRACTION")
            logger.info("=" * 80)

            extracted_urls = []

            # Extract hyperlinks if openpyxl is available
            if HAS_OPENPYXL:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.hyperlink and cell.hyperlink.target:
                                    url = cell.hyperlink.target
                                    if url not in extracted_urls:
                                        extracted_urls.append(url)
                    logger.info(f"Extracted {len(extracted_urls)} hyperlinks from Excel file")
                except Exception as e:
                    logger.error(f"Error extracting hyperlinks with openpyxl: {e}")

            # Convert to markdown (no character limits for comprehensive extraction)
            markdown_content = self.spreadsheet_to_markdown(content_bytes, filename)

            logger.info("=" * 80)
            logger.info(f"COMPREHENSIVE MARKDOWN EXTRACTION COMPLETE:")
            logger.info(f"  - Total content: {len(markdown_content)} characters")
            logger.info(f"  - Hyperlinks extracted: {len(extracted_urls)}")
            logger.info("=" * 80)

            # Log extracted URLs
            if extracted_urls:
                logger.info("EXTRACTED HYPERLINKS:")
                for i, url in enumerate(extracted_urls[:10]):
                    logger.info(f"  URL {i + 1}: {url}")
                if len(extracted_urls) > 10:
                    logger.info(f"  ... and {len(extracted_urls) - 10} more URLs")

            return markdown_content, extracted_urls

        except Exception as e:
            logger.error(f"Error extracting comprehensive Excel content as Markdown: {e}")
            # Fallback to basic markdown conversion
            markdown_content = self.spreadsheet_to_markdown(content_bytes, filename)
            return markdown_content, []