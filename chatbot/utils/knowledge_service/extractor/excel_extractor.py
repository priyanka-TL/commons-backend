"""Excel extraction functionality"""

import io
import logging
from typing import List, Dict, Any, Tuple
import pandas as pd

logger = logging.getLogger('django')

try:
    import openpyxl

    HAS_OPENPYXL = True
    logger.info("openpyxl available for enhanced Excel processing")
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not available. Excel hyperlink extraction will be limited.")


class ExcelExtractor:
    """Handles Excel content extraction"""

    def __init__(self, excel_max_rows: int = 50, excel_max_cols: int = 20,
                 subdoc_max_chars: int = 500):
        self.excel_max_rows = excel_max_rows
        self.excel_max_cols = excel_max_cols
        self.subdoc_max_chars = subdoc_max_chars

    def extract_limited_content(self, content_bytes: bytes, max_chars: int = None) -> str:
        """Extract limited content from Excel file for LLM processing

        Args:
            content_bytes: Excel file content as bytes
            max_chars: Maximum characters to extract

        Returns:
            Limited text content for LLM processing
        """
        if max_chars is None:
            max_chars = self.subdoc_max_chars

        try:
            excel_file = pd.ExcelFile(io.BytesIO(content_bytes))
            sheet_names = excel_file.sheet_names

            logger.info("=" * 80)
            logger.info(f"EXCEL FILE CONTAINS {len(sheet_names)} SHEETS:")
            for i, sheet_name in enumerate(sheet_names):
                logger.info(f"  Sheet {i + 1}: '{sheet_name}'")
            logger.info("=" * 80)

            if not sheet_names:
                return ""

            # Process sheets until we have enough content
            all_text_parts = []
            total_chars = 0
            sheets_processed = 0

            for sheet_idx, sheet_name in enumerate(sheet_names):
                if total_chars >= max_chars:
                    break

                logger.info(f"Processing sheet {sheet_idx + 1}: '{sheet_name}'")

                try:
                    # Read the sheet
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name
                    )

                    # Skip empty sheets
                    if df.empty or len(df) == 0:
                        logger.warning(f"Sheet '{sheet_name}' is empty, moving to next sheet...")
                        continue

                    # Limit rows and columns for processing
                    display_df = df.head(self.excel_max_rows)
                    if len(df.columns) > self.excel_max_cols:
                        display_df = display_df.iloc[:, :self.excel_max_cols]

                    # Convert to CSV-like format
                    csv_string = display_df.to_csv(index=False)

                    # Add sheet header if we're processing multiple sheets
                    if sheets_processed > 0:
                        all_text_parts.append(f"\n\n--- Sheet: '{sheet_name}' ---\n")

                    all_text_parts.append(csv_string)
                    sheets_processed += 1

                    # Update total characters
                    current_text = '\n'.join(all_text_parts)
                    total_chars = len(current_text)

                    logger.info(f"Sheet '{sheet_name}' added {len(csv_string)} chars (total: {total_chars} chars)")

                    # Add truncation note for this sheet if needed
                    if len(df) > self.excel_max_rows or len(df.columns) > self.excel_max_cols:
                        all_text_parts.append(
                            f"\n[Sheet '{sheet_name}': Showing {min(len(df), self.excel_max_rows)} of {len(df)} rows, "
                            f"{min(len(df.columns), self.excel_max_cols)} of {len(df.columns)} columns]"
                        )

                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {e}")
                    continue

            # If no sheets had data
            if sheets_processed == 0:
                logger.warning("All sheets are empty!")
                return "All Excel sheets are empty (no data found)"

            # Join all parts
            full_text = '\n'.join(all_text_parts)
            original_length = len(full_text)

            logger.info(f"Processed {sheets_processed} sheets with data, extracted {original_length} chars")

            # Log the content
            logger.info("=" * 80)
            logger.info("EXCEL CONTENT BEING SENT TO LLM:")
            logger.info("=" * 80)
            logger.info(full_text)
            logger.info("=" * 80)

            # Apply final character limit if needed
            if len(full_text) > max_chars:
                # Try to cut at a row boundary
                lines = full_text.split('\n')
                truncated_text = []
                current_length = 0

                for line in lines:
                    if current_length + len(line) + 1 > max_chars - 50:
                        break
                    truncated_text.append(line)
                    current_length += len(line) + 1

                full_text = '\n'.join(truncated_text) + "\n[Content truncated]"
                logger.info(f"Excel content truncated from {original_length} to {len(full_text)} chars")

            return full_text

        except Exception as e:
            logger.error(f"Error extracting Excel content: {e}")
            return ""

    def extract_comprehensive_content_for_urls(self, content_bytes: bytes) -> Tuple[str, List[str]]:
        """Extract COMPLETE Excel content from ALL sheets and hyperlinks using openpyxl

        Args:
            content_bytes: Excel file content as bytes

        Returns:
            Tuple of (comprehensive_text, extracted_urls)
        """
        try:
            logger.info("=" * 80)
            logger.info("EXTRACTING COMPREHENSIVE EXCEL CONTENT FOR URL EXTRACTION (OPENPYXL)")
            logger.info("=" * 80)

            if not HAS_OPENPYXL:
                logger.warning("openpyxl not available, falling back to pandas method")
                return self._extract_full_content_for_urls(content_bytes), []

            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            sheet_names = wb.sheetnames

            logger.info(f"Processing ALL {len(sheet_names)} sheets for content and URL extraction:")
            for i, sheet_name in enumerate(sheet_names):
                logger.info(f"  Sheet {i + 1}: '{sheet_name}'")

            all_text_parts = []
            extracted_urls = []
            total_urls_found = 0

            # Process ALL sheets without any limits
            for sheet_idx, sheet_name in enumerate(sheet_names):
                logger.info(
                    f"Processing sheet {sheet_idx + 1}/{len(sheet_names)}: '{sheet_name}' for content and URLs...")

                try:
                    sheet = wb[sheet_name]

                    # Check if sheet has data
                    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
                        logger.info(f"  Sheet '{sheet_name}' is empty, skipping...")
                        continue

                    logger.info(f"  Sheet '{sheet_name}': {sheet.max_row} rows x {sheet.max_column} columns")

                    # Extract content in multiple formats
                    sheet_text_parts = []
                    sheet_urls = []

                    # Add sheet header
                    sheet_text_parts.append(f"\n=== SHEET: {sheet_name} ===")

                    # Extract column headers (first row)
                    headers = []
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(1, col)
                        header_value = cell.value
                        if header_value is not None:
                            headers.append(str(header_value))
                        else:
                            headers.append(f"Unnamed: {col - 1}")

                    sheet_text_parts.append("COLUMNS: " + " | ".join(headers))

                    # Process each row
                    for row_idx in range(1, sheet.max_row + 1):
                        row_content = []
                        row_has_content = False

                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row_idx, col_idx)

                            # Extract hyperlinks
                            if cell.hyperlink and cell.hyperlink.target:
                                url = cell.hyperlink.target
                                if url not in sheet_urls:
                                    sheet_urls.append(url)
                                    extracted_urls.append(url)

                            # Extract cell content
                            cell_value = cell.value
                            if cell_value is not None:
                                cell_str = str(cell_value).strip()
                                if cell_str:
                                    col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"Col{col_idx}"
                                    row_content.append(f"{col_name}: {cell_str}")
                                    row_has_content = True

                        if row_has_content:
                            sheet_text_parts.append(f"ROW {row_idx}: " + " | ".join(row_content))

                    # Also add CSV-like format for compatibility
                    sheet_text_parts.append("\n--- CSV FORMAT ---")
                    csv_rows = []
                    for row_idx in range(1, sheet.max_row + 1):
                        csv_row = []
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row_idx, col_idx)
                            cell_value = cell.value
                            if cell_value is not None:
                                csv_row.append(str(cell_value))
                            else:
                                csv_row.append("")
                        csv_rows.append(",".join(f'"{item}"' for item in csv_row))

                    sheet_text_parts.extend(csv_rows)

                    # Join all parts for this sheet
                    sheet_content = '\n'.join(sheet_text_parts)

                    # Count URLs in this sheet for logging
                    sheet_url_count = len(sheet_urls)
                    total_urls_found += sheet_url_count

                    logger.info(
                        f"  Sheet '{sheet_name}' content: {len(sheet_content)} chars, {sheet_url_count} hyperlinks extracted")

                    all_text_parts.append(sheet_content)

                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}' with openpyxl: {e}")
                    continue

            # Join all sheet content
            complete_content = '\n\n'.join(all_text_parts)

            logger.info("=" * 80)
            logger.info(f"COMPREHENSIVE EXCEL EXTRACTION COMPLETE (OPENPYXL):")
            logger.info(f"  - Processed {len(sheet_names)} sheets")
            logger.info(f"  - Total content: {len(complete_content)} characters")
            logger.info(f"  - Hyperlinks extracted: {len(extracted_urls)}")
            logger.info(f"  - Total URLs found: {total_urls_found}")
            logger.info("=" * 80)

            # Log extracted URLs
            if extracted_urls:
                logger.info("EXTRACTED HYPERLINKS:")
                for i, url in enumerate(extracted_urls[:10]):  # Log first 10
                    logger.info(f"  URL {i + 1}: {url}")
                if len(extracted_urls) > 10:
                    logger.info(f"  ... and {len(extracted_urls) - 10} more URLs")

            # Log sample content
            sample_content = complete_content[:2000] if len(complete_content) > 2000 else complete_content
            logger.info("SAMPLE OF COMPREHENSIVE EXCEL CONTENT:")
            logger.info(sample_content)
            if len(complete_content) > 2000:
                logger.info(f"... [TRUNCATED - FULL CONTENT IS {len(complete_content)} CHARS] ...")
            logger.info("=" * 80)

            return complete_content, extracted_urls

        except Exception as e:
            logger.error(f"Error extracting comprehensive Excel content with openpyxl: {e}")
            # Fallback to pandas method
            return self._extract_full_content_for_urls(content_bytes), []

    def _extract_full_content_for_urls(self, content_bytes: bytes) -> str:
        """Fallback: Extract full Excel content specifically for URL extraction - no limits

        Args:
            content_bytes: Excel file content as bytes

        Returns:
            Full text content for URL extraction
        """
        try:
            excel_file = pd.ExcelFile(io.BytesIO(content_bytes))
            sheet_names = excel_file.sheet_names

            all_text_parts = []

            # Process ALL sheets without limits
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    if not df.empty:
                        # Convert entire sheet to string
                        csv_string = df.to_csv(index=False)
                        all_text_parts.append(f"\n--- Sheet: '{sheet_name}' ---\n")
                        all_text_parts.append(csv_string)
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {e}")
                    continue

            return '\n'.join(all_text_parts)

        except Exception as e:
            logger.error(f"Error extracting full Excel content: {e}")
            return ""

    def extract_text(self, file_path) -> str:
        """Extract text from Excel (file path) - first sheet only with limits

        Args:
            file_path: Path to Excel file

        Returns:
            Limited text content
        """
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        if not sheet_names:
            return ""

        # Only read first sheet
        df = pd.read_excel(excel_file, sheet_name=sheet_names[0], nrows=self.excel_max_rows)
        if len(df.columns) > self.excel_max_cols:
            df = df.iloc[:, :self.excel_max_cols]

        text = f"Excel file with {len(sheet_names)} sheets. Processing first sheet: '{sheet_names[0]}'\n"
        text += df.to_string(max_rows=self.excel_max_rows, max_cols=self.excel_max_cols)
        return text

    def extract_text_from_object(self, file) -> str:
        """Extract text from Excel (file object) - first sheet only with limits

        Args:
            file: Excel file object

        Returns:
            Limited text content
        """
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names

        if not sheet_names:
            return ""

        # Only read first sheet
        df = pd.read_excel(excel_file, sheet_name=sheet_names[0], nrows=self.excel_max_rows)
        if len(df.columns) > self.excel_max_cols:
            df = df.iloc[:, :self.excel_max_cols]

        text = f"Excel file with {len(sheet_names)} sheets. Processing first sheet: '{sheet_names[0]}'\n"
        text += df.to_string(max_rows=self.excel_max_rows, max_cols=self.excel_max_cols)
        return text